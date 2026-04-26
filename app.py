"""
🧠 Student Cognitive Performance Report Generator
A beautiful, parent-friendly visualization tool for tracking student progress.
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from datetime import datetime, timedelta
import os
import io
import gspread
from google.oauth2.service_account import Credentials

# ============================================================================
# CONFIGURATION
# ============================================================================

COGNITIVE_STACKS = [
    "Attention & Focus",
    "Pattern Recognition",
    "Problem Solving",
    "Executive Function",
    "Emotional Regulation",
    "Metacognition"
]

# Parent-friendly descriptions
STACK_DESCRIPTIONS = {
    "Attention & Focus": "Staying focused during activities",
    "Pattern Recognition": "Spotting patterns and visual thinking",
    "Problem Solving": "Problem-solving and critical thinking",
    "Executive Function": "Planning, organizing, and completing tasks",
    "Emotional Regulation": "Managing emotions and staying calm",
    "Metacognition": "Self-awareness and learning from mistakes"
}

# Beautiful color palette
COLORS = {
    'excellent': '#2ECC71',    # Green
    'good': '#3498DB',         # Blue
    'developing': '#F39C12',   # Orange
    'needs_attention': '#E74C3C',  # Red
    'primary': '#2E86AB',
    'secondary': '#A23B72',
    'background': '#F8F9FA',
    'text': '#2C3E50'
}

DATA_FILE = "student_data.csv"
DATA_FILE_XLSX = "student_data.xlsx"

# ============================================================================
# PAGE CONFIGURATION
# ============================================================================

st.set_page_config(
    page_title="Cognitive Performance Tracker",
    page_icon="🧠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for beautiful styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #2E86AB;
        text-align: center;
        padding: 1rem 0;
        margin-bottom: 2rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-top: -1rem;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 1rem;
        color: white;
        text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .score-excellent { color: #2ECC71; font-weight: bold; }
    .score-good { color: #3498DB; font-weight: bold; }
    .score-developing { color: #F39C12; font-weight: bold; }
    .score-needs-attention { color: #E74C3C; font-weight: bold; }
    .stButton > button {
        width: 100%;
        border-radius: 0.5rem;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# DATA MANAGEMENT — Google Sheets with local CSV fallback
# ============================================================================

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

def _get_gsheet_connection():
    """Return (spreadsheet, True) if Google Sheets is configured, else (None, False)."""
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
        sheet_url = st.secrets["spreadsheet_url"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_url(sheet_url)
        try:
            spreadsheet.worksheet("student_data")
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title="student_data", rows=1000, cols=10)
            ws.append_row(['Student', 'Date'] + COGNITIVE_STACKS)
        return spreadsheet, True
    except Exception:
        return None, False

COLUMN_RENAMES = {
    "Logical Reasoning": "Problem Solving",
}

def _migrate_columns(df):
    """Rename legacy column names to current ones."""
    for old_name, new_name in COLUMN_RENAMES.items():
        if old_name in df.columns and new_name not in df.columns:
            df = df.rename(columns={old_name: new_name})
    return df

_SOURCE_COL = "_source_tab"

def load_data():
    """Load student data from ALL tabs in Google Sheets (or local CSV as fallback).

    Every worksheet in the spreadsheet is read and merged into one DataFrame.
    Each row is tagged with a hidden ``_source_tab`` column so save_data()
    knows which rows belong to the app's own tab vs. instructor tabs.
    Tabs that don't contain the expected columns are silently skipped.
    """
    spreadsheet, using_sheets = _get_gsheet_connection()
    expected_cols = {'Student', 'Date'} | set(COGNITIVE_STACKS)

    if using_sheets:
        try:
            all_frames = []
            for ws in spreadsheet.worksheets():
                try:
                    records = ws.get_all_records()
                    if not records:
                        continue
                    tab_df = pd.DataFrame(records)
                    tab_df = _migrate_columns(tab_df)
                    if not expected_cols.issubset(set(tab_df.columns)):
                        missing = expected_cols - set(tab_df.columns)
                        st.sidebar.caption(
                            f"⚠️ Tab '{ws.title}' skipped — missing: {missing}"
                        )
                        continue
                    tab_df = tab_df[['Student', 'Date'] + COGNITIVE_STACKS]
                    tab_df[_SOURCE_COL] = ws.title
                    all_frames.append(tab_df)
                except Exception as tab_err:
                    st.sidebar.caption(f"⚠️ Tab '{ws.title}': {tab_err}")
                    continue

            if all_frames:
                df = pd.concat(all_frames, ignore_index=True)
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                return df
            return pd.DataFrame(columns=['Student', 'Date'] + COGNITIVE_STACKS + [_SOURCE_COL])
        except Exception as e:
            st.error(f"Error loading from Google Sheets: {e}")
            return pd.DataFrame(columns=['Student', 'Date'] + COGNITIVE_STACKS + [_SOURCE_COL])
    else:
        if os.path.exists(DATA_FILE_XLSX):
            try:
                all_frames = []
                xls = pd.ExcelFile(DATA_FILE_XLSX)
                for sheet_name in xls.sheet_names:
                    try:
                        tab_df = pd.read_excel(xls, sheet_name=sheet_name)
                        tab_df = _migrate_columns(tab_df)
                        if not expected_cols.issubset(set(tab_df.columns)):
                            continue
                        tab_df = tab_df[['Student', 'Date'] + COGNITIVE_STACKS]
                        tab_df[_SOURCE_COL] = sheet_name
                        all_frames.append(tab_df)
                    except Exception:
                        continue
                if all_frames:
                    df = pd.concat(all_frames, ignore_index=True)
                    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                    return df
                return pd.DataFrame(columns=['Student', 'Date'] + COGNITIVE_STACKS + [_SOURCE_COL])
            except Exception as e:
                st.error(f"Error loading Excel file: {e}")
                return pd.DataFrame(columns=['Student', 'Date'] + COGNITIVE_STACKS + [_SOURCE_COL])
        elif os.path.exists(DATA_FILE):
            try:
                df = pd.read_csv(DATA_FILE)
                df = _migrate_columns(df)
                if _SOURCE_COL not in df.columns:
                    df[_SOURCE_COL] = "student_data"
                if len(df) > 0 and 'Date' in df.columns:
                    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                return df
            except Exception as e:
                st.error(f"Error loading data: {e}")
                return pd.DataFrame(columns=['Student', 'Date'] + COGNITIVE_STACKS + [_SOURCE_COL])
        return pd.DataFrame(columns=['Student', 'Date'] + COGNITIVE_STACKS + [_SOURCE_COL])

def save_data(df):
    """Save only app-owned rows to the 'student_data' tab (or local CSV).

    Rows that came from instructor tabs are left untouched — only rows
    tagged as 'student_data' are written back.
    """
    spreadsheet, using_sheets = _get_gsheet_connection()

    if using_sheets:
        try:
            ws = spreadsheet.worksheet("student_data")
            app_df = df[df[_SOURCE_COL] == "student_data"].copy()
            save_df = app_df.drop(columns=[_SOURCE_COL])
            save_df['Date'] = save_df['Date'].apply(
                lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else ''
            )
            ws.clear()
            ws.append_row(save_df.columns.tolist())
            if len(save_df) > 0:
                ws.append_rows(save_df.values.tolist())
        except Exception as e:
            st.error(f"Error saving to Google Sheets: {e}")
    else:
        app_df = df[df[_SOURCE_COL] == "student_data"].copy()
        save_df = app_df.drop(columns=[_SOURCE_COL])
        if os.path.exists(DATA_FILE_XLSX):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(DATA_FILE_XLSX)
                if "student_data" not in wb.sheetnames:
                    wb.create_sheet("student_data")
                ws = wb["student_data"]
                ws.delete_rows(1, ws.max_row)
                for r_idx, row in enumerate([save_df.columns.tolist()] + save_df.values.tolist()):
                    for c_idx, val in enumerate(row):
                        ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
                wb.save(DATA_FILE_XLSX)
            except Exception as e:
                st.error(f"Error saving to Excel: {e}")
        else:
            save_df.to_csv(DATA_FILE, index=False)

def get_students(df):
    """Get list of unique students"""
    if len(df) > 0 and 'Student' in df.columns:
        return sorted(df['Student'].dropna().unique().tolist())
    return []

def add_score(df, student, date, scores):
    """Add or update a score entry in the app-owned tab only.

    If the same student+date already exists in the 'student_data' tab it is
    replaced. Rows from instructor tabs are never touched.
    """
    if len(df) > 0 and 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        try:
            dup_mask = (
                (df['Student'] == student)
                & (df['Date'].dt.date == date)
                & (df[_SOURCE_COL] == "student_data")
            )
            df = df[~dup_mask].copy()
        except Exception:
            df = df.copy()
    else:
        df = df.copy()

    new_row = {'Student': student, 'Date': pd.Timestamp(date), _SOURCE_COL: "student_data"}
    new_row.update(scores)
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

    return df.sort_values(['Student', 'Date'])

# ============================================================================
# VISUALIZATION FUNCTIONS (Parent-Friendly)
# ============================================================================

def get_score_color(score):
    """Get color from a smooth red-orange-yellow-green gradient based on score"""
    cmap = LinearSegmentedColormap.from_list(
        'score_gradient', ['#E74C3C', '#E67E22', '#F1C40F', '#2ECC71']
    )
    normalized = np.clip((score - 1) / 9, 0, 1)
    return cmap(normalized)

def get_score_label(score):
    """Get parent-friendly label for score"""
    if score >= 8:
        return "Excellent! ⭐"
    elif score >= 6:
        return "Good Progress 👍"
    elif score >= 4:
        return "Developing 📈"
    else:
        return "Needs Support 💪"


def get_tier(score):
    """Classify a score into one of four tiers."""
    if score >= 8:
        return "excellent"
    elif score >= 6:
        return "good"
    elif score >= 4:
        return "developing"
    else:
        return "needs_support"


def _is_high(score):
    return score >= 6


def generate_personalized_message(scores, student_name):
    """
    Build a 2-sentence personalized message from layered templates.

    Sentence 1 — Cognitive chain: Attention → Pattern Rec → Problem Solving → Exec Function
    Sentence 2 — Emotional Regulation + Metacognition

    Scores dict keys must match COGNITIVE_STACKS names.
    """
    af = scores["Attention & Focus"]
    pr = scores["Pattern Recognition"]
    ps = scores["Problem Solving"]
    ef = scores["Executive Function"]
    er = scores["Emotional Regulation"]
    mc = scores["Metacognition"]

    chain = [af, pr, ps, ef]
    chain_high = [_is_high(s) for s in chain]

    # --- Sentence 1: cognitive chain ---
    if all(chain_high):
        s1 = (
            f"{student_name} shows strong focus, recognises patterns from previous games, "
            f"and uses them to solve problems and plan ahead confidently."
        )
    elif chain_high[0] and chain_high[1] and chain_high[2] and not chain_high[3]:
        s1 = (
            f"{student_name} focuses well, spots patterns, and solves problems independently "
            f"— the next step is learning to plan further ahead and think through full strategies."
        )
    elif chain_high[0] and chain_high[1] and not chain_high[2]:
        s1 = (
            f"{student_name} concentrates well and is getting better at recognising ideas from "
            f"past games — the next exciting step is using those observations to solve positions independently."
        )
    elif chain_high[0] and not chain_high[1]:
        s1 = (
            f"{student_name} is able to stay focused during sessions, which is a great foundation "
            f"— the next step is learning to spot familiar patterns and ideas from previous games."
        )
    elif not any(chain_high):
        s1 = (
            f"{student_name} is building the important foundations right now — staying focused and "
            f"spotting patterns are the first big steps, and these will unlock problem solving and "
            f"thinking ahead over time."
        )
    else:
        best_chain_labels = ["staying focused", "spotting patterns",
                             "solving problems", "planning ahead"]
        strongest_idx = int(np.argmax(chain))
        weakest_idx = int(np.argmin(chain))
        s1 = (
            f"{student_name} shows promising ability in {best_chain_labels[strongest_idx]}, "
            f"and building on {best_chain_labels[weakest_idx]} will help bring everything together."
        )

    # --- Sentence 2: emotional regulation + metacognition ---
    er_high = _is_high(er)
    mc_high = _is_high(mc)

    if er_high and mc_high:
        s2 = (
            f"{'She' if student_name.endswith('a') or student_name.endswith('i') else 'He'} "
            f"stays calm under pressure and explains thinking well — keep up the great work!"
        )
    elif er_high and not mc_high:
        s2 = (
            f"{'She' if student_name.endswith('a') or student_name.endswith('i') else 'He'} "
            f"handles pressure well at the board, and encouraging "
            f"{'her' if student_name.endswith('a') or student_name.endswith('i') else 'him'} "
            f"to talk about why {'she' if student_name.endswith('a') or student_name.endswith('i') else 'he'} "
            f"made certain moves will help develop even further."
        )
    elif not er_high and mc_high:
        s2 = (
            f"{'She' if student_name.endswith('a') or student_name.endswith('i') else 'He'} "
            f"explains reasoning behind moves well, and encouraging "
            f"{'her' if student_name.endswith('a') or student_name.endswith('i') else 'him'} "
            f"to stay positive after tough games will help put that thinking to work consistently."
        )
    else:
        s2 = (
            f"Encouraging {student_name} to stay positive after tough moments and talk through "
            f"why {'she' if student_name.endswith('a') or student_name.endswith('i') else 'he'} "
            f"made certain moves will help everything come together."
        )

    return f"{s1} {s2}"

def create_progress_snapshot(student_data, student_name):
    """Create a bar chart with a personalized message block below it."""
    import textwrap

    student_data = student_data.copy()
    student_data['Date'] = pd.to_datetime(student_data['Date'], errors='coerce')

    fig, (ax_chart, ax_msg) = plt.subplots(
        2, 1, figsize=(12, 10), facecolor='white',
        gridspec_kw={'height_ratios': [3, 1], 'hspace': 0.15}
    )

    if len(student_data) > 1:
        display_scores = student_data[COGNITIVE_STACKS].mean().values
        first_scores = student_data[COGNITIVE_STACKS].iloc[0].values
        last_scores = student_data[COGNITIVE_STACKS].iloc[-1].values
        improvements = last_scores - first_scores
    else:
        display_scores = student_data[COGNITIVE_STACKS].iloc[-1].values
        improvements = np.zeros(len(COGNITIVE_STACKS))

    min_date = student_data['Date'].min()
    max_date = student_data['Date'].max()
    fig.suptitle(f"✨ {student_name}'s Learning Journey ✨",
                 fontsize=22, fontweight='bold', color=COLORS['text'], y=0.97)
    if len(student_data) > 1:
        date_range = f"{min_date.strftime('%d %b')} - {max_date.strftime('%d %b %Y')}"
        fig.text(0.5, 0.92, f"Average Scores • {date_range} ({len(student_data)} sessions)",
                 fontsize=11, ha='center', color='gray')
    else:
        fig.text(0.5, 0.92, f"Session on {max_date.strftime('%d %b %Y')}",
                 fontsize=11, ha='center', color='gray')

    # --- Bar chart (top subplot) ---
    ax = ax_chart
    cmap = LinearSegmentedColormap.from_list(
        'score_gradient', ['#E74C3C', '#E67E22', '#F1C40F', '#2ECC71']
    )
    n_stacks = len(COGNITIVE_STACKS)
    y_pos = np.arange(n_stacks)
    bar_height = 0.7
    half_h = bar_height / 2
    bar_max = 10

    ax.set_xlim(-0.2, 12.5)
    ax.set_ylim(n_stacks - 0.5, -0.5)
    ax.set_autoscale_on(False)
    ax.set_facecolor('#FAFAFA')
    chart_title = "Average Scores at a Glance" if len(student_data) > 1 else "Skills at a Glance"
    ax.set_title(chart_title, fontsize=16, fontweight='bold', pad=15)
    ax.set_xlabel('Score', fontsize=11)
    ax.set_xticks(range(0, 11, 2))
    ax.set_yticks(y_pos)
    ax.set_yticklabels(COGNITIVE_STACKS, fontsize=11)

    for i, (score, improvement) in enumerate(zip(display_scores, improvements)):
        bg = plt.Rectangle((0, i - half_h), bar_max, bar_height,
                           facecolor='white', edgecolor='none', zorder=1)
        ax.add_patch(bg)

        if score > 0:
            gradient = np.linspace(0, score / bar_max, 256).reshape(1, -1)
            ax.imshow(gradient, aspect='auto', cmap=cmap, vmin=0, vmax=1,
                      extent=[0, score, i + half_h, i - half_h],
                      zorder=2, interpolation='bilinear')

        c = get_score_color(score)
        text_color = (c[0] * 0.7, c[1] * 0.7, c[2] * 0.7, 1.0)
        ax.text(bar_max + 0.3, i, f'{score:.1f}/10', va='center',
                fontsize=11, fontweight='bold', color=text_color, zorder=4)

        if len(student_data) > 1:
            if improvement > 0:
                ax.text(bar_max + 1.3, i, f'↑{int(improvement)}', va='center',
                        fontsize=10, color='green', fontweight='bold', zorder=4)
            elif improvement < 0:
                ax.text(bar_max + 1.3, i, f'↓{int(abs(improvement))}', va='center',
                        fontsize=10, color='red', fontweight='bold', zorder=4)

    # --- Summary bar ---
    avg_score = np.mean(display_scores)
    total_improvement = np.sum(improvements) if len(student_data) > 1 else 0

    summary = f"📊 Overall Score: {avg_score:.1f}/10"
    if len(student_data) > 1:
        summary += f"  |  📈 Total Growth: {'+' if total_improvement >= 0 else ''}{total_improvement:.0f} points"
    summary += f"  |  📅 Sessions: {len(student_data)}"

    # --- Personalized message (bottom subplot) ---
    score_dict = {stack: float(s) for stack, s in zip(COGNITIVE_STACKS, display_scores)}
    message = generate_personalized_message(score_dict, student_name)

    ax_msg.set_axis_off()
    wrapped = textwrap.fill(message, width=95)

    ax_msg.text(
        0.5, 0.82, "Instructor's Note",
        transform=ax_msg.transAxes, fontsize=14, fontweight='bold',
        ha='center', va='top', color=COLORS['primary']
    )
    ax_msg.text(
        0.5, 0.62, wrapped,
        transform=ax_msg.transAxes, fontsize=11.5, ha='center', va='top',
        color=COLORS['text'], linespacing=1.5,
        fontstyle='italic'
    )
    ax_msg.add_patch(plt.Rectangle(
        (0.03, 0.05), 0.94, 0.9, transform=ax_msg.transAxes,
        facecolor='#F0F7FB', edgecolor=COLORS['primary'],
        linewidth=1.5, zorder=0, clip_on=False
    ))

    fig.text(0.5, 0.02, summary, fontsize=10, ha='center', color=COLORS['text'],
             bbox=dict(boxstyle='round,pad=0.5', facecolor='#E8F4FD',
                       edgecolor=COLORS['primary'], alpha=0.95))

    plt.tight_layout(rect=[0, 0.05, 1, 0.90])

    return fig



# ============================================================================
# STREAMLIT APP
# ============================================================================

def main():
    # Load data
    if 'data' not in st.session_state:
        st.session_state.data = load_data()
    
    df = st.session_state.data
    
    # Header
    st.markdown('<h1 class="main-header">🧠 Cognitive Performance Tracker</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Track and visualize student cognitive development</p>', unsafe_allow_html=True)
    
    # Sidebar navigation
    st.sidebar.title("📋 Menu")
    page = st.sidebar.radio("Navigate to:", 
                           ["📝 Enter Scores", "📊 Generate Report", "👥 Manage Students", "📁 View All Data"])
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📈 Quick Stats")
    students = get_students(df)
    st.sidebar.metric("Total Students", len(students))
    st.sidebar.metric("Total Records", len(df))
    
    # ==================== PAGE: ENTER SCORES ====================
    if page == "📝 Enter Scores":
        st.header("📝 Enter Daily Scores")
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("Student Selection")
            
            # Option to add new or select existing
            input_mode = st.radio("", ["Select existing student", "Add new student"], horizontal=True)
            
            if input_mode == "Add new student":
                student_name = st.text_input("Enter student name:", placeholder="e.g., Rahul Sharma")
            else:
                if students:
                    student_name = st.selectbox("Select student:", students)
                else:
                    st.info("No students yet. Add a new student above!")
                    student_name = None
            
            # Date selection
            score_date = st.date_input("📅 Date:", value=datetime.today())
        
        with col2:
            st.subheader("Enter Scores (1-10)")
            
            if student_name:
                scores = {}
                
                for stack in COGNITIVE_STACKS:
                    col_a, col_b = st.columns([3, 1])
                    with col_a:
                        st.markdown(f"**{stack}**")
                        st.caption(STACK_DESCRIPTIONS[stack])
                    with col_b:
                        scores[stack] = st.number_input(
                            stack, min_value=1, max_value=10, value=5,
                            key=f"score_{stack}", label_visibility="collapsed"
                        )
                
                st.markdown("---")
                
                if st.button("💾 Save Scores", type="primary", use_container_width=True):
                    if student_name.strip():
                        st.session_state.data = add_score(df, student_name.strip(), score_date, scores)
                        save_data(st.session_state.data)
                        st.success(f"✅ Scores saved for {student_name} on {score_date.strftime('%d %b %Y')}!")
                        st.balloons()
                    else:
                        st.error("Please enter a student name.")
    
    # ==================== PAGE: GENERATE REPORT ====================
    elif page == "📊 Generate Report":
        st.header("📊 Generate Visual Report")
        
        if not students:
            st.warning("No student data yet. Please enter some scores first!")
            return
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.subheader("Report Settings")
            
            selected_student = st.selectbox("👤 Select Student:", students)
            
            st.markdown("---")
            st.markdown("**📅 Date Range:**")
            
            date_option = st.radio("", [
                "All time",
                "Last 7 days",
                "Last 30 days",
                "Custom range"
            ])
            
            # Get student's data date range
            student_df = df[df['Student'] == selected_student].copy()
            # Ensure Date is datetime
            student_df['Date'] = pd.to_datetime(student_df['Date'], errors='coerce')
            student_df = student_df.dropna(subset=['Date'])
            
            if len(student_df) == 0:
                st.warning("No valid date entries for this student.")
                return
            
            min_date = student_df['Date'].min().date()
            max_date = student_df['Date'].max().date()
            
            if date_option == "Custom range":
                date_from = st.date_input("From:", value=min_date, min_value=min_date, max_value=max_date)
                date_to = st.date_input("To:", value=max_date, min_value=min_date, max_value=max_date)
            elif date_option == "Last 7 days":
                date_to = max_date
                date_from = max(min_date, max_date - timedelta(days=7))
            elif date_option == "Last 30 days":
                date_to = max_date
                date_from = max(min_date, max_date - timedelta(days=30))
            else:  # All time
                date_from = min_date
                date_to = max_date
            
            st.markdown("---")
            
            generate_btn = st.button("🎨 Generate Report", type="primary", use_container_width=True)
        
        with col2:
            if generate_btn:
                # Filter data by date range (Date is already datetime from above)
                mask = (student_df['Date'].dt.date >= date_from) & (student_df['Date'].dt.date <= date_to)
                filtered_df = student_df[mask].sort_values('Date')
                
                if len(filtered_df) == 0:
                    st.error("No data found for the selected date range.")
                else:
                    st.subheader(f"Report for {selected_student}")
                    
                    with st.spinner("Creating beautiful report..."):
                        fig = create_progress_snapshot(filtered_df, selected_student)
                        
                        st.pyplot(fig)
                        
                        # Download button
                        buf = io.BytesIO()
                        fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', 
                                   facecolor='white', edgecolor='none')
                        buf.seek(0)
                        
                        filename = f"Report_{selected_student.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.png"
                        
                        st.download_button(
                            label="⬇️ Download Report (PNG)",
                            data=buf,
                            file_name=filename,
                            mime="image/png",
                            use_container_width=True
                        )
                        
                        plt.close(fig)
    
    # ==================== PAGE: MANAGE STUDENTS ====================
    elif page == "👥 Manage Students":
        st.header("👥 Manage Students")
        
        if not students:
            st.info("No students yet. Go to 'Enter Scores' to add students.")
            return
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("📋 Student List")
            for i, student in enumerate(students, 1):
                student_records = len(df[df['Student'] == student])
                st.markdown(f"**{i}. {student}** ({student_records} records)")
        
        with col2:
            st.subheader("🗑️ Remove Student")
            st.warning("⚠️ This will delete ALL data for the selected student!")
            
            student_to_remove = st.selectbox("Select student to remove:", students)
            
            if st.button("🗑️ Remove Student", type="secondary"):
                app_mask = (df['Student'] == student_to_remove) & (df[_SOURCE_COL] == "student_data")
                st.session_state.data = df[~app_mask]
                save_data(st.session_state.data)
                st.success(f"Removed {student_to_remove} from app data")
                st.rerun()
    
    # ==================== PAGE: VIEW ALL DATA ====================
    elif page == "📁 View All Data":
        st.header("📁 All Student Data")
        
        if len(df) == 0:
            st.info("No data yet. Start by entering scores!")
        else:
            # Filter by student
            filter_student = st.selectbox("Filter by student:", ["All Students"] + students)
            
            if filter_student == "All Students":
                display_df = df.copy()
            else:
                display_df = df[df['Student'] == filter_student].copy()
            
            display_df = display_df.drop(columns=[_SOURCE_COL], errors='ignore')

            # Format display - ensure Date is datetime first
            display_df['Date'] = pd.to_datetime(display_df['Date'], errors='coerce')
            display_df['Date'] = display_df['Date'].apply(
                lambda x: x.strftime('%d %b %Y') if pd.notna(x) else ''
            )
            
            st.dataframe(display_df, use_container_width=True, hide_index=True)
            
            # Download CSV
            csv = display_df.to_csv(index=False)
            st.download_button(
                "⬇️ Download as CSV",
                csv,
                "student_data_export.csv",
                "text/csv",
                use_container_width=True
            )

if __name__ == "__main__":
    main()
